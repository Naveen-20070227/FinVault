import csv
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO, StringIO
import calendar
from typing import Dict, Any, List, Tuple

from sqlalchemy import func
from sqlalchemy.orm import Session

# Excel & PDF imports
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from app.models.transaction import Transaction
from app.models.category import Category
from app.repositories.transaction_repo import TransactionRepository
from app.repositories.category_repo import CategoryRepository
from app.repositories.user_repo import UserRepository

class ReportService:
    def __init__(self, db: Session):
        self.db = db
        self.tx_repo = TransactionRepository(db)
        self.cat_repo = CategoryRepository(db)
        self.user_repo = UserRepository(db)

    def _get_month_range(self, month_str: str) -> Tuple[date, date, int, int]:
        try:
            year, m = map(int, month_str.split("-"))
        except ValueError:
            today = date.today()
            year, m = today.year, today.month
            
        start_date = date(year, m, 1)
        if m == 12:
            end_date = date(year + 1, 1, 1)
        else:
            end_date = date(year, m + 1, 1)
            
        return start_date, end_date, m, year

    def get_monthly_report_data(self, user_id: int, month_str: str) -> Dict[str, Any]:
        start_date, end_date, month, year = self._get_month_range(month_str)
        user = self.user_repo.get_by_id(user_id)
        currency = user.currency if user else "₹"
        
        # 1. Totals
        total_income = self.db.query(func.sum(Transaction.amount)).filter(
            Transaction.user_id == user_id,
            Transaction.type == "income",
            Transaction.date >= start_date,
            Transaction.date < end_date
        ).scalar() or Decimal("0.00")

        total_expense = self.db.query(func.sum(Transaction.amount)).filter(
            Transaction.user_id == user_id,
            Transaction.type == "expense",
            Transaction.date >= start_date,
            Transaction.date < end_date
        ).scalar() or Decimal("0.00")

        net_savings = total_income - total_expense
        savings_rate = float((net_savings / total_income) * 100) if total_income > 0 else 0.0

        # 2. Avg daily spend
        num_days = calendar.monthrange(year, month)[1]
        today = date.today()
        # If calculating for current month, use current day for active average, else total month days
        days_in_calc = today.day if (today.month == month and today.year == year) else num_days
        avg_daily_spend = total_expense / Decimal(str(days_in_calc)) if days_in_calc > 0 else Decimal("0.00")

        # 3. Category Breakdown
        spent_grouped = self.db.query(
            Transaction.category_id,
            func.sum(Transaction.amount).label("total_spent")
        ).filter(
            Transaction.user_id == user_id,
            Transaction.type == "expense",
            Transaction.date >= start_date,
            Transaction.date < end_date
        ).group_by(Transaction.category_id).all()

        breakdown = []
        for cat_id, spent in spent_grouped:
            cat = self.cat_repo.get_by_id(cat_id)
            if cat:
                pct = float((spent / total_expense) * 100) if total_expense > 0 else 0.0
                breakdown.append({
                    "category_name": cat.name,
                    "amount": spent,
                    "percentage": pct,
                    "color": cat.color
                })
        breakdown.sort(key=lambda x: x["amount"], reverse=True)

        return {
            "month_label": start_date.strftime("%B %Y"),
            "currency": currency,
            "total_income": total_income,
            "total_expense": total_expense,
            "net_savings": net_savings,
            "savings_rate": savings_rate,
            "avg_daily_spend": avg_daily_spend,
            "category_breakdown": breakdown
        }

    def get_yearly_report_data(self, user_id: int, year: int) -> Dict[str, Any]:
        user = self.user_repo.get_by_id(user_id)
        currency = user.currency if user else "₹"
        
        monthly_table = []
        total_income = Decimal("0.00")
        total_expense = Decimal("0.00")

        for month in range(1, 13):
            start_date, end_date, _, _ = self._get_month_range(f"{year}-{month:02d}")
            
            inc = self.db.query(func.sum(Transaction.amount)).filter(
                Transaction.user_id == user_id,
                Transaction.type == "income",
                Transaction.date >= start_date,
                Transaction.date < end_date
            ).scalar() or Decimal("0.00")

            exp = self.db.query(func.sum(Transaction.amount)).filter(
                Transaction.user_id == user_id,
                Transaction.type == "expense",
                Transaction.date >= start_date,
                Transaction.date < end_date
            ).scalar() or Decimal("0.00")

            savings = inc - exp
            total_income += inc
            total_expense += exp

            monthly_table.append({
                "month_num": month,
                "month_name": calendar.month_name[month],
                "income": inc,
                "expense": exp,
                "savings": savings
            })

        net_savings = total_income - total_expense
        savings_rate = float((net_savings / total_income) * 100) if total_income > 0 else 0.0

        return {
            "year": year,
            "currency": currency,
            "total_income": total_income,
            "total_expense": total_expense,
            "net_savings": net_savings,
            "savings_rate": savings_rate,
            "monthly_table": monthly_table
        }

    def export_csv(self, user_id: int, month_str: str) -> StringIO:
        start_date, end_date, _, _ = self._get_month_range(month_str)
        transactions = self.db.query(Transaction).filter(
            Transaction.user_id == user_id,
            Transaction.date >= start_date,
            Transaction.date < end_date
        ).order_by(Transaction.date.desc()).all()

        output = StringIO()
        writer = csv.writer(output)
        
        # Header
        writer.writerow(["Date", "Title", "Type", "Category", "Amount", "Notes"])
        
        for tx in transactions:
            cat = self.cat_repo.get_by_id(tx.category_id)
            cat_name = cat.name if cat else "Uncategorized"
            writer.writerow([
                tx.date.strftime("%Y-%m-%d"),
                tx.title,
                tx.type.capitalize(),
                cat_name,
                float(tx.amount),
                tx.notes or ""
            ])
            
        output.seek(0)
        return output

    def export_xlsx(self, user_id: int, month_str: str) -> BytesIO:
        start_date, end_date, _, _ = self._get_month_range(month_str)
        transactions = self.db.query(Transaction).filter(
            Transaction.user_id == user_id,
            Transaction.date >= start_date,
            Transaction.date < end_date
        ).order_by(Transaction.date.desc()).all()

        user = self.user_repo.get_by_id(user_id)
        currency = user.currency if user else "₹"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transactions"
        ws.views.sheetView[0].showGridLines = True

        # Styles
        purple_theme_color = "7C3AED"
        accent_light = "F3E8FF"
        
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color=purple_theme_color, end_color=purple_theme_color, fill_type="solid")
        title_font = Font(name="Segoe UI", size=16, bold=True, color=purple_theme_color)
        regular_font = Font(name="Segoe UI", size=10)
        total_font = Font(name="Segoe UI", size=11, bold=True)
        
        thin_border = Border(
            left=Side(style="thin", color="E5E7EB"),
            right=Side(style="thin", color="E5E7EB"),
            top=Side(style="thin", color="E5E7EB"),
            bottom=Side(style="thin", color="E5E7EB")
        )
        double_bottom_border = Border(
            top=Side(style="thin", color="111111"),
            bottom=Side(style="double", color="111111")
        )

        # Title Block
        ws["A1"] = f"FinVault Transaction Report — {start_date.strftime('%B %Y')}"
        ws["A1"].font = title_font
        ws.row_dimensions[1].height = 30
        
        # Metadata Block
        ws["A2"] = f"Exported on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws["A2"].font = Font(name="Segoe UI", size=9, italic=True)
        ws["A3"] = f"Currency: {currency}"
        ws["A3"].font = Font(name="Segoe UI", size=9, italic=True)
        
        # Headers Row
        headers = ["Date", "Title", "Type", "Category", "Amount", "Notes"]
        ws.row_dimensions[5].height = 24
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Fill Data
        row_idx = 6
        for tx in transactions:
            cat = self.cat_repo.get_by_id(tx.category_id)
            cat_name = cat.name if cat else "Uncategorized"
            
            ws.cell(row=row_idx, column=1, value=tx.date).number_format = "yyyy-mm-dd"
            ws.cell(row=row_idx, column=2, value=tx.title)
            ws.cell(row=row_idx, column=3, value=tx.type.capitalize())
            ws.cell(row=row_idx, column=4, value=cat_name)
            
            amt_cell = ws.cell(row=row_idx, column=5, value=float(tx.amount))
            amt_cell.number_format = f'"{currency}"#,##0.00'
            
            ws.cell(row=row_idx, column=6, value=tx.notes or "")

            for col_num in range(1, 7):
                c = ws.cell(row=row_idx, column=col_num)
                c.font = regular_font
                c.border = thin_border
                if col_num in [1, 3]:
                    c.alignment = Alignment(horizontal="center")
                elif col_num == 5:
                    c.alignment = Alignment(horizontal="right")
                    
            ws.row_dimensions[row_idx].height = 20
            row_idx += 1

        # Summary / Totals block
        ws.cell(row=row_idx, column=4, value="Total Spending").font = total_font
        ws.cell(row=row_idx, column=4).alignment = Alignment(horizontal="right")
        
        # Formula for sum of expenses
        spent_sum_cell = ws.cell(
            row=row_idx,
            column=5,
            value=f'=SUMIF(C6:C{row_idx-1}, "Expense", E6:E{row_idx-1})'
        )
        spent_sum_cell.font = total_font
        spent_sum_cell.number_format = f'"{currency}"#,##0.00'
        spent_sum_cell.border = double_bottom_border

        # Adjust columns widths
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            # Skip title row for length check
            for cell in col[4:]:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["F"].width = 30

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def export_pdf(self, user_id: int, month_str: str) -> bytes:
        data = self.get_monthly_report_data(user_id, month_str)
        start_date, end_date, _, _ = self._get_month_range(month_str)
        
        transactions = self.db.query(Transaction).filter(
            Transaction.user_id == user_id,
            Transaction.date >= start_date,
            Transaction.date < end_date
        ).order_by(Transaction.date.desc()).all()

        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            rightMargin=36,
            leftMargin=36,
            topMargin=36,
            bottomMargin=36
        )

        styles = getSampleStyleSheet()
        
        # Color definitions matching purple theme
        primary_color = colors.HexColor("#7C3AED")
        text_dark = colors.HexColor("#1A0533")
        light_purple = colors.HexColor("#F3E8FF")
        border_grey = colors.HexColor("#E5E7EB")

        # Custom paragraph styles
        title_style = ParagraphStyle(
            "PDFTitle",
            parent=styles["Title"],
            fontName="Helvetica-Bold",
            fontSize=22,
            textColor=primary_color,
            alignment=0,
            spaceAfter=6
        )
        subtitle_style = ParagraphStyle(
            "PDFSubtitle",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=10,
            textColor=colors.HexColor("#6B4D8A"),
            spaceAfter=20
        )
        section_heading = ParagraphStyle(
            "PDFSection",
            parent=styles["Heading2"],
            fontName="Helvetica-Bold",
            fontSize=14,
            textColor=text_dark,
            spaceBefore=15,
            spaceAfter=8
        )
        cell_bold = ParagraphStyle(
            "PDFCellBold",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=9,
            textColor=text_dark
        )
        cell_regular = ParagraphStyle(
            "PDFCellReg",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=9,
            textColor=text_dark
        )

        elements = []

        # 1. Header
        elements.append(Paragraph(f"FinVault Monthly Report", title_style))
        elements.append(Paragraph(f"Statement Period: {data['month_label']} | Generated on: {datetime.now().strftime('%Y-%m-%d')}", subtitle_style))
        
        # 2. KPI Summary Cards Grid
        summary_data = [
            [
                Paragraph("<b>Total Income</b>", cell_bold),
                Paragraph("<b>Total Expenses</b>", cell_bold),
                Paragraph("<b>Net Savings</b>", cell_bold),
                Paragraph("<b>Savings Rate</b>", cell_bold)
            ],
            [
                Paragraph(f"{data['currency']}{data['total_income']:,.2f}", cell_regular),
                Paragraph(f"{data['currency']}{data['total_expense']:,.2f}", cell_regular),
                Paragraph(f"{data['currency']}{data['net_savings']:,.2f}", cell_regular),
                Paragraph(f"{data['savings_rate']:.1f}%", cell_regular)
            ]
        ]
        
        summary_table = Table(summary_data, colWidths=[1.8*inch, 1.8*inch, 1.8*inch, 1.8*inch])
        summary_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), light_purple),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 8),
            ("BOX", (0, 0), (-1, -1), 1, primary_color),
            ("INNERGRID", (0, 0), (-1, -1), 0.5, border_grey),
        ]))
        
        elements.append(summary_table)
        elements.append(Spacer(1, 15))

        # 3. Category Breakdown Table
        elements.append(Paragraph("Category Spending Breakdown", section_heading))
        
        cat_rows = [[Paragraph("<b>Category</b>", cell_bold), Paragraph("<b>Amount Spent</b>", cell_bold), Paragraph("<b>Percentage</b>", cell_bold)]]
        for item in data["category_breakdown"]:
            cat_rows.append([
                Paragraph(item["category_name"], cell_regular),
                Paragraph(f"{data['currency']}{item['amount']:,.2f}", cell_regular),
                Paragraph(f"{item['percentage']:.1f}%", cell_regular)
            ])
            
        if len(cat_rows) == 1:
            cat_rows.append([Paragraph("No expense transactions in this month", cell_regular), Paragraph("-", cell_regular), Paragraph("-", cell_regular)])

        cat_table = Table(cat_rows, colWidths=[3.2*inch, 2.0*inch, 2.0*inch])
        cat_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F9F9FB")),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("LINEBELOW", (0, 0), (-1, -1), 0.5, border_grey),
            ("LINEBELOW", (0, 0), (-1, 0), 1.5, primary_color),
        ]))
        
        elements.append(cat_table)
        elements.append(Spacer(1, 15))

        # 4. Recent Transactions Table
        elements.append(Paragraph("Transaction History", section_heading))
        
        tx_rows = [[
            Paragraph("<b>Date</b>", cell_bold),
            Paragraph("<b>Title</b>", cell_bold),
            Paragraph("<b>Type</b>", cell_bold),
            Paragraph("<b>Category</b>", cell_bold),
            Paragraph("<b>Amount</b>", cell_bold)
        ]]
        
        for tx in transactions[:100]:  # Cap at 100 for PDF layout safety
            cat = self.cat_repo.get_by_id(tx.category_id)
            cat_name = cat.name if cat else "Uncategorized"
            color_type = colors.HexColor("#10B981") if tx.type == "income" else colors.HexColor("#EF4444")
            
            tx_rows.append([
                Paragraph(tx.date.strftime("%Y-%m-%d"), cell_regular),
                Paragraph(tx.title, cell_regular),
                Paragraph(f"<font color='{color_type}'><b>{tx.type.capitalize()}</b></font>", cell_regular),
                Paragraph(cat_name, cell_regular),
                Paragraph(f"<b>{data['currency']}{tx.amount:,.2f}</b>", cell_regular)
            ])

        if len(tx_rows) == 1:
            tx_rows.append([Paragraph("No transactions recorded for this period", cell_regular), Paragraph("-", cell_regular), Paragraph("-", cell_regular), Paragraph("-", cell_regular), Paragraph("-", cell_regular)])

        tx_table = Table(tx_rows, colWidths=[1.1*inch, 2.4*inch, 1.0*inch, 1.4*inch, 1.3*inch])
        tx_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F9F9FB")),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("LINEBELOW", (0, 0), (-1, -1), 0.5, border_grey),
            ("LINEBELOW", (0, 0), (-1, 0), 1.5, primary_color),
        ]))
        
        elements.append(tx_table)

        # Build Document
        doc.build(elements)
        pdf_bytes = buffer.getvalue()
        buffer.close()
        return pdf_bytes
